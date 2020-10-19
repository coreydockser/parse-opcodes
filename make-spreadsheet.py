#!/usr/bin/env python
# coding: utf-8

# In[170]:


import pandas as pd
import numpy as np
from os import path
import re
import math
from openpyxl import load_workbook


# In[171]:


with open("opcode-formats", "r") as f:
        opcode_formats = [line.rstrip() for line in f]
    


# In[172]:


#splits it into argluts and opcodes
#specifically, takes the first instance of a title and then the first empty line after the title 
#and splices the chosen list using those indices

# ensures that this function will still work even if the last item is not an empty string
if opcode_formats[-1] is not '':
    opcode_formats.append('')
    

arg_idx = opcode_formats.index('##ARGLUTS')
arg_first_empty = opcode_formats[arg_idx + 1:].index('')
arglut_list = opcode_formats[arg_idx + 1 : arg_first_empty]

opcode_idx = opcode_formats.index('##OPCODES')
opcode_first_empty = opcode_formats[opcode_idx + 1:].index('')
opcodes = opcode_formats[opcode_idx + 1 : opcode_first_empty]


# In[173]:


#turns the opcodes list into a list of lists.
#each item in opcodes_list is a list of the items in that particular row, to make it easier to parse
opcodes_list = []
for item in opcodes:
    temp_list = re.sub(' +', ' ', item).split()
    opcodes_list.append(temp_list)
opcodes_list


# In[174]:


## executes the items in arglut, so that they become a dict
for item in arglut_list:
    exec(item)


# In[176]:


## parses opcodes into text
## and opcode consists of a digit, followed by two periods a second digit, and equal sign, and a hex number
def opcode_to_binary(opcode):
    opcode_split = re.split("\.\.|\=", opcode)
    
    #Code to convert hex to binary
    if opcode_split[2] != 'ignore':    
        opcode_split[2] = bin(int(opcode_split[2], 16))[2:]
    return opcode_split


# In[177]:


#splits args into a list of numbers for the purposes of assigning bit in the final worksheet
def arg_to_binary(arg):
    split_arg = re.split("\[|\||\]", arg)[1:-1] # the split is to keep imm out
    temp_arg_list = []
    for entry in split_arg:
        if ":" in entry:
            entry_split = re.split("\:", entry)
            entry_range = range(int(entry_split[0]), int(entry_split[1]) - 1, -1)
            for i in entry_range:
                temp_arg_list.append(i)
        else:
            temp_arg_list.append(int(entry))
    return temp_arg_list
        
    


# parses formatted args into lists of the values they represent, for use later


arglut_values = {}
def format_argluts():
    argluts_formatted = list(arglut_format.values())
    for arg in argluts_formatted:
        arglut_values[arg] = arg_to_binary(arg)
        
    
format_argluts()


# In[179]:


## a function which takes a list of strings (opcodes) and parses them, creating a list of 32 items 
## corresponding to the 32 bits available
numtup = ('0', '1', '2', '3', '4', '5', '6', '7', '8', '9')
def opcode_to_bits(opcode_set):
## first step is to find out how many args are in the set
    opcode_dict = {}
    opcode_dict_sorted = {}
## first item in the list is just the name of the instruction. it takes  up the first bit, #31
    opcode_dict['instr'] = opcode_set[0]
##  each following item has  a unique place based on the instruction given to it
    for entry in opcode_set[1:]:
        ## function that places  things in based on the area
        ## if the line starts with a string its an opcode
        if entry.startswith(numtup):
            ## translated is the opcode separated into a list of three digits
            ## for example, 14..12=0 becomes [14, 12, 0]
            translated = opcode_to_binary(entry)
            first = int(translated[0])
            last = int(translated[1])
            opcode_len = first - last + 1
            opcode_dict[first] = translated[2].zfill(opcode_len)
            for num in range(last, first):
                opcode_dict[num] = '%'
            
        ## otherwise its an arg
        ## args have their range contained in a tuple in arglut. this code just extracts that
        else:
            first = arglut[entry][0]
            last = arglut[entry][1]
            opcode_dict[first] = entry
            for num in range(last, first):
                opcode_dict[num] = '%'
    
    ## this then sorts it in reverse order,  manually since i cant sort by key
    opcode_dict_sorted['instr'] = opcode_dict['instr']
    for i in range(31, -1, -1):
        opcode_dict_sorted[i] = opcode_dict[i]
        

    return opcode_dict_sorted


# In[180]:


# takes a list of opcodes and runs opcode_to_bits() on each entry, then appends it to a separate list
def opcodes_bits_to_lists(opcode_list):
    dict_list = []
    for entry in opcode_list:
        thing = opcode_to_bits(entry)
        dict_list.append(thing)
    return dict_list


# In[182]:


opcodes_dict_list = opcodes_bits_to_lists(opcodes_list)


# In[183]:


df = pd.DataFrame(data= opcodes_dict_list) 
cols = ['instr']
for n in range(31, -1, -1):
    cols.append(str(n))

df.columns=cols


# In[185]:


#replaces args with the correct words without disrupting anything
df =  df.replace(arglut_format)


# In[186]:


## generate two sheets, one is formatted with merge and center, while the other has names like imm[20:15]31 or whatever
with pd.ExcelWriter('opcodes.xlsx') as writer:  
    df.to_excel(writer, sheet_name='with_merge')
    df.to_excel(writer, sheet_name='without_merge')


# # Part 2 -> Formatting the Worksheet

# In[188]:


#loads the workbook
wb = load_workbook('opcodes.xlsx')


# In[189]:


print(wb.sheetnames)


# In[190]:


from openpyxl.styles import Alignment

with_merge=wb['with_merge']
without_merge=wb['without_merge']


# In[191]:


row_count=with_merge.max_row
column_count=with_merge.max_column


# In[192]:



#works in a row r by going down the list, keeping track of a column#, c. it checks to see if c's  nextdoor neighbor    
#a recursive function that
def page_through_withm(c, r):
   ws = with_merge
   curr_cell = ws.cell(column=c, row=r)
   next_cell = ws.cell(column=curr_cell.column + 1, row =r)
   length_count = 0
   
   #checks to see if the current cell is a %, if it is then something has gone wrong
   if curr_cell.value != '%':
       curr_cell.alignment = Alignment(horizontal='center')
       # checks  to see if the next cell is a %, if it is then a merge is needed, otherwise continue on
       while next_cell.value == '%':
           length_count += 1
           next_cell = ws.cell(column=next_cell.column + 1, row = r)
       dest_c = curr_cell.column + length_count
       # determines what to do: merge (if it starts with a letter), or split (if it starts with a number)?
       if curr_cell.value.startswith(numtup):
           stored_value = curr_cell.value # stores the value of curr_cell 
           stored_column = curr_cell.column
           for i in range(length_count + 1):
               ws.cell(column=stored_column + i, row=r).value = stored_value[i]
               ws.cell(column=stored_column + i, row=r).alignment = Alignment(horizontal='center')
       else:
           with_merge.merge_cells(start_row=r, start_column=c, end_row=r, end_column=dest_c)
       #loops so long as dest_c + 1, which is the cell after the merge, isnt the last
       if dest_c + 1 < 34:
           page_through_withm(dest_c + 1, r)
   #loops if the next cell isnt out of bounds       
   elif c + 1 < 34:
       page_through_withm(c + 1, r)
       
       
   else:
       print(curr_cell.value)
       print("Error: Cannot pass cell with % into page_through_withm") 
   


# In[193]:


# a tuple with the values in arglut_format, for checking to see if it needs specially formatting
argtup = tuple(arglut_format.values())

## basically identical to the above except it formats the args differently
def page_through_withoutm(c, r):
    wsm = without_merge
    curr_cell = wsm.cell(column=c, row=r)
    next_cell = wsm.cell(column=curr_cell.column + 1, row = r)
    length_count = 0
    
    
    #checks to see if the current cell is a %, if it is then something has gone wrong
    if curr_cell.value != '%':
        curr_cell.alignment = Alignment(horizontal='center')
        # checks  to see if the next cell is a %, if it is then a merge is needed, otherwise continue on
        while next_cell.value == '%':
            length_count += 1
            next_cell = wsm.cell(column=next_cell.column + 1, row = r)
        dest_c = curr_cell.column + length_count
        # determines what to do: merge (if it starts with a letter), or split (if it starts with a number)?
        if curr_cell.value.startswith(numtup):
            stored_value = curr_cell.value # stores the value of curr_cell 
            stored_column = curr_cell.column
            for i in range(length_count + 1):
                wsm.cell(column=stored_column + i, row=r).value = stored_value[i]
                wsm.cell(column=stored_column + i, row=r).alignment = Alignment(horizontal='center')
        else:
            ## if the current cell value is in argtup 
            if curr_cell.value in argtup:
                stored_value = curr_cell.value
                stored_column = curr_cell.column
                stored_tup = arglut_values[curr_cell.value]
                for i in range(length_count + 1):
                    wsm.cell(column = stored_column + i, row = r).value = 'imm[' + str(stored_tup[i]) + "]"
                    wsm.cell(column=stored_column + i, row=r).alignment = Alignment(horizontal='center')
            else:
                without_merge.merge_cells(start_row=r, start_column=c, end_row=r, end_column=dest_c)
        #loops so long as dest_c + 1, which is the cell after the merge, isnt the last
        if dest_c + 1 < 34:
            page_through_withoutm(dest_c + 1, r)
    #loops if the next cell isnt out of bounds       
    elif c + 1 < 34:
        page_through_withoutm(c + 1, r)
        
        
    else:
        print(curr_cell.value)
        print("Error: Cannot pass cell with % into page_through_withm") 


# In[194]:


for r in range(2, row_count + 1): #we're skipping a row so it starts at 1
    page_through_withm(3, r)
    


# In[195]:


for r in range(2, row_count + 1):
    page_through_withoutm(3, r)


# In[196]:


#saves the workbook, finalizing the changes
wb.save('opcodes.xlsx')

